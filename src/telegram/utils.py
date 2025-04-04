from pathlib import Path
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def create_excel(data: List[Dict], save_path: Path, date_start: str, date_end: str):
    wb = Workbook()
    ws = wb.active

    headers = list(data[0].keys())
    ws.append(headers)

    for entry in data:
        row = [entry[key] for key in headers]
        ws.append(row)

    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for row in data:
            value = str(row[headers[col - 1]])
            max_length = max(max_length, len(value))
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    file_path = save_path / f"expenses_{date_start}-{date_end}.xlsx"
    wb.save(file_path)
    return file_path
